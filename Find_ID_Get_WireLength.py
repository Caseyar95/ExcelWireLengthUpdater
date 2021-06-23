import openpyxl as xl

#opening workbook, loading sheetnames
filename ="4206RSheets.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]
ws2 = wb1.worksheets[1]

# calculate total number of rows
mr = ws1.max_row
mr2 = ws2.max_row

for y in range (1, mr2 + 1):
    v = ws2.cell(row = y, column = 3).value
    #print (v) ###TEST
    for i in range (1, mr + 1):
        c = ws1.cell(row = i, column = 3).value
        #print (c) ###TEST
        if (c == v):
            ws2.cell(row = y, column = 7).value = ws1.cell(row = i, column = 7).value

#saving file
wb1.save(str(filename))
